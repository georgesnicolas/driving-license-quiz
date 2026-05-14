"""Extract quiz questions from the Arabic xlsx into questions.json + images folder."""
import json
import os
import re
from pathlib import Path

import openpyxl
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor

SRC = Path(r"C:\Users\georges.nicolas\Downloads\Code arabic 21.xlsx")
OUT_DIR = Path(__file__).parent
IMG_DIR = OUT_DIR / "images"
IMG_DIR.mkdir(exist_ok=True)


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active

    # ---- Save images and capture anchor rows ----
    image_by_row = {}  # 1-indexed spreadsheet row -> image filename
    for idx, img in enumerate(ws._images, start=1):
        anchor = img.anchor
        if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
            row = anchor._from.row + 1  # 0-indexed -> 1-indexed
        else:
            continue

        # Determine extension
        ref = getattr(img, "path", None) or getattr(img, "ref", "")
        ext = ".jpeg"
        if isinstance(ref, str):
            m = re.search(r"\.(jpe?g|png|gif|bmp)$", ref, re.I)
            if m:
                ext = "." + m.group(1).lower()
        # Fetch raw bytes
        data = None
        for attr in ("_data", "data"):
            d = getattr(img, attr, None)
            if callable(d):
                try:
                    data = d()
                    break
                except Exception:
                    pass
            elif d is not None:
                data = d
                break
        if data is None:
            # fall back to reading from zip via img.ref
            continue
        fname = f"img_{idx:03d}{ext}"
        (IMG_DIR / fname).write_bytes(data)
        image_by_row.setdefault(row, []).append(fname)

    # ---- Parse questions ----
    # Layout: every 3 consecutive rows starting at row 3 form one question.
    # Column A (1) = answer text; bold => correct answer.
    # Column B (2) = question text, merged across the 3 rows (may be empty for image-only questions).
    # Column E (5) = category; Column H (8) = question number (only on first row, sometimes empty).
    questions = []
    BLOCK = 3
    r = 3
    def first_nonempty(row, cols):
        for c in cols:
            v = ws.cell(row, c).value
            if v is not None and str(v).strip():
                return v
        return None

    while r + BLOCK - 1 <= ws.max_row:
        # Layout shifts across the sheet: question may be in cols B-D,
        # category in cols E-G, number in cols H-J.
        q_text = first_nonempty(r, (2, 3, 4))
        category = first_nonempty(r, (5, 6, 7))
        number = first_nonempty(r, (8, 9, 10))
        answers = []
        for i in range(BLOCK):
            cell = ws.cell(r + i, 1)
            v = cell.value
            if v is None or not str(v).strip():
                continue
            answers.append({
                "text": str(v).strip(),
                "correct": bool(cell.font and cell.font.b),
            })
        img_files = image_by_row.get(r, [])
        image = img_files[0] if img_files else None

        # Skip completely empty blocks
        if not answers and not (q_text and str(q_text).strip()) and not image:
            r += BLOCK
            continue

        questions.append({
            "number": number,
            "category": str(category).strip() if category else None,
            "question": str(q_text).strip() if q_text else None,
            "answers": answers,
            "image": image,
        })
        r += BLOCK

    # Ensure each question has exactly one correct answer (warn otherwise)
    no_correct = sum(1 for q in questions if not any(a["correct"] for a in q["answers"]))
    multi = sum(1 for q in questions if sum(1 for a in q["answers"] if a["correct"]) > 1)
    print(f"Total questions: {len(questions)}")
    print(f"Without correct answer: {no_correct}")
    print(f"With multiple correct: {multi}")
    print(f"Images saved: {sum(len(v) for v in image_by_row.values())}")
    print(f"Questions with image: {sum(1 for q in questions if q['image'])}")

    (OUT_DIR / "questions.json").write_text(
        json.dumps(questions, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print("Wrote questions.json")


if __name__ == "__main__":
    main()
